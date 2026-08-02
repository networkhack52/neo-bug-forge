"""Write approved answers back into an .xlsx the customer can return.

Two modes:
  * ``export_simple`` — a clean two-column (Question / Answer) workbook,
    always available.
Answers are written next to their questions with a confidence column so the
reviewer can spot anything still flagged.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def export_simple(name: str, items: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Responses"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["#", "Question", "Answer", "Confidence", "Status"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    widths = [5, 55, 70, 12, 12]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + c)].width = w

    review_fill = PatternFill("solid", fgColor="FEF3C7")
    for i, item in enumerate(items, start=1):
        r = i + 1
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=item["question"]).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=3, value=item.get("answer", "")).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=4, value=round(float(item.get("confidence", 0)), 1))
        status = "review" if item.get("needs_review") else "ok"
        scell = ws.cell(row=r, column=5, value=status)
        if item.get("needs_review"):
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = review_fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
