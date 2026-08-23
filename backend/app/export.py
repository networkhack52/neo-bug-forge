"""Write approved answers back into an .xlsx the customer can return.

Two modes:
  * ``export_simple`` — a clean two-column (Question / Answer) workbook,
    always available.
  * ``export_original`` — the customer's own uploaded file with the status
    and answer cells filled in, so they get their exact template back.
Answers are written next to their questions with a confidence column so the
reviewer can spot anything still flagged.
"""
from __future__ import annotations

import csv
import io
import json
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Internal review/status notes (e.g. "[Attestly review: …]") are for the review
# screen only — strip them from anything the customer receives.
_NOTE_RE = re.compile(r"\s*\[Attestly[^\]]*\]", re.IGNORECASE)

LOCKED_TEXT = "Locked · upgrade to answer"


def clean_answer(text: str) -> str:
    return _NOTE_RE.sub("", text or "").strip()


def _citations(item: dict) -> list[dict]:
    raw = item.get("citations")
    if isinstance(raw, list):
        return raw
    try:
        return json.loads(raw or "[]")
    except (ValueError, TypeError):
        return []


def status_of(item: dict) -> str:
    """One of Answered / Needs review / No evidence / N/A / Locked.

    'No evidence' (your documents don't say) is kept distinct from a negative
    answer of 'No' (you don't do this) — the latter reports as Answered and
    carries a remediation note in the response text."""
    if item.get("locked"):
        return "Locked"
    if item.get("excluded") or (item.get("choice") or "").strip() == "Not Applicable":
        return "N/A"
    if item.get("needs_review"):
        return "Needs review" if _citations(item) else "No evidence"
    return "Answered"


def _negative_suffix(item: dict) -> str:
    """The remediation note a 'No' must carry so it never exports bare."""
    date = str(item.get("remediation_date") or "").strip()
    if date:
        return f" Remediation planned by {date}."
    if item.get("no_plan"):
        return " No remediation is currently planned."
    return ""


def is_bare_negative(item: dict) -> bool:
    """True when a 'No' answer is genuinely bare — essentially just the word 'No'
    with no explanation. A substantive negative (e.g. 'We do not use customer data
    to train models…') is a complete answer and is NOT a control gap, so it does
    NOT need a remediation date."""
    if (item.get("choice") or "").strip() != "No":
        return False
    ans = clean_answer(item.get("answer", "")).strip().lower()
    ans = re.sub(r"^(no|yes|not applicable|n/?a)[\s.:,;-]*", "", ans).strip()
    return len(ans) < 12


def gate_issues(items: list[dict]) -> list[dict]:
    """Rows that can't export yet: a BARE 'No' (just 'No' with no explanation and
    no remediation date / no-plan note) or an 'N/A' with no justification. A
    substantive 'No' that explains itself exports freely.
    Returns [{id, question, issue}] so the caller can prompt for what's missing."""
    out = []
    for it in items:
        if it.get("locked"):
            continue
        choice = (it.get("choice") or "").strip()
        if it.get("excluded"):
            if not str(it.get("exclusion_reason") or "").strip():
                out.append({"id": it.get("id"), "question": it.get("question", ""),
                            "issue": "na_reason"})
            continue
        if choice == "No" and is_bare_negative(it):
            if not str(it.get("remediation_date") or "").strip() and not it.get("no_plan"):
                out.append({"id": it.get("id"), "question": it.get("question", ""),
                            "issue": "remediation"})
        elif choice == "Not Applicable":
            if not str(it.get("na_reason") or "").strip():
                out.append({"id": it.get("id"), "question": it.get("question", ""),
                            "issue": "na_reason"})
    return out


def _date_suffix(c: dict) -> str:
    """`` (reviewed 2025-03-14)`` for a dated document citation, adding a
    staleness note when the source is past the freshness threshold. Empty for
    undated or library citations."""
    date = str(c.get("date") or "").strip()
    if not date:
        return ""
    word = "reviewed" if c.get("date_basis") == "stated" else "dated"
    note = ", review recommended" if c.get("stale") else ""
    return f" ({word} {date}{note})"


def source_cell(item: dict) -> str:
    """Document name + quoted line per citation, one per line, with the source
    date appended so a reviewer can see how fresh the evidence is."""
    lines = []
    for c in _citations(item):
        title = str(c.get("title") or "").strip()
        text = str(c.get("text") or "").strip()
        label = title if c.get("kind") == "document" else (title or "Approved answer")
        suffix = _date_suffix(c)
        if label and text:
            lines.append(f'{label}: "{text}"{suffix}')
        elif label:
            lines.append(f"{label}{suffix}")
        elif text:
            lines.append(f'"{text}"{suffix}')
    return "\n".join(lines)


def vendor_response(item: dict) -> str:
    """The answer text the customer sends: status prefix + detail, cleaned.

    Never emits a bare 'No' or 'N/A' — a negative carries its remediation note
    and an N/A carries its justification (the triage exclusion reason when the
    row was excluded, else the user's one-line reason)."""
    if item.get("locked"):
        return LOCKED_TEXT
    choice = (item.get("choice") or "").strip()
    # Excluded (out-of-scope) rows export as N/A with the exclusion reason.
    if item.get("excluded"):
        reason = str(item.get("exclusion_reason") or "").strip()
        return f"Not Applicable. {reason}" if reason else "Not Applicable."
    answer = clean_answer(item.get("answer", ""))
    if choice == "Not Applicable":
        reason = str(item.get("na_reason") or "").strip()
        body = answer if answer and answer.lower() != "not applicable" else ""
        detail = " ".join(x for x in (body, reason) if x)
        return f"Not Applicable. {detail}".strip() if detail else "Not Applicable."
    base = f"{choice}. {answer}" if (choice and answer and not answer.lower().startswith(choice.lower())) else (answer or choice)
    if choice == "No":
        base = (base + _negative_suffix(item)).strip()
    return base


def export_simple(name: str, items: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Responses"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    # The whole product is "proof of answer": every response carries its Source
    # (document + quoted line) and a Status the reviewer can scan.
    headers = ["Section", "Question", "Vendor Response", "Source", "Status"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    widths = [18, 55, 60, 50, 14]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + c)].width = w

    review_fill = PatternFill("solid", fgColor="FEF3C7")
    locked_fill = PatternFill("solid", fgColor="E5E7EB")
    wrap_top = Alignment(wrap_text=True, vertical="top")
    for i, item in enumerate(items, start=1):
        r = i + 1
        status = status_of(item)
        ws.cell(row=r, column=1, value=item.get("section", "")).alignment = wrap_top
        ws.cell(row=r, column=2, value=item["question"]).alignment = wrap_top
        ws.cell(row=r, column=3, value=vendor_response(item)).alignment = wrap_top
        ws.cell(row=r, column=4, value=source_cell(item)).alignment = wrap_top
        ws.cell(row=r, column=5, value=status).alignment = Alignment(vertical="top")
        fill = locked_fill if status == "Locked" else (review_fill if item.get("needs_review") else None)
        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _original_cells(item: dict, has_detail: bool) -> tuple[str, str]:
    """Split an item into (status_cell, detail_cell) text for the customer's own
    template. Honours the same negative/N-A rules as `vendor_response`.

    When the template has a distinct comments column the status goes in the
    answer column and the free-text in the comments column. Otherwise the two
    are combined into the single answer column.
    """
    excluded = bool(item.get("excluded"))
    choice = (item.get("choice") or "").strip()
    if excluded or choice == "Not Applicable":
        status = "Not Applicable"
        reason = str(item.get("exclusion_reason") if excluded else item.get("na_reason") or "").strip()
        body = "" if excluded else clean_answer(item.get("answer", ""))
        detail = " ".join(x for x in (body if body.lower() != "not applicable" else "", reason) if x)
        if has_detail:
            return status, detail
        return (f"{status}. {detail}".strip() if detail else status), ""

    answer = clean_answer(item.get("answer", ""))
    if choice == "No":
        answer = (answer + _negative_suffix(item)).strip()
    if has_detail:
        return choice, answer
    if choice and answer:
        if answer.lower().startswith(choice.lower()):
            return answer, ""
        return f"{choice}. {answer}", ""
    return (choice or answer), ""


def can_export_original(q: dict) -> bool:
    """True when we captured enough of the upload to fill it back in."""
    return bool(q.get("source_bytes")) and bool(q.get("answer_col"))


def export_original(q: dict, items: list[dict]) -> bytes:
    """Fill the customer's own uploaded file and return the same format."""
    source = q.get("source_bytes")
    if not source:
        raise ValueError("no original file stored for this questionnaire")
    answer_col = q.get("answer_col")
    detail_col = q.get("detail_col")
    has_detail = bool(detail_col)
    kind = (q.get("source_kind") or "xlsx").lower()

    if kind == "csv":
        return _fill_csv(source, items, answer_col, detail_col, has_detail)
    return _fill_xlsx(source, q.get("sheet_name"), items, answer_col, detail_col, has_detail)


def _fill_xlsx(source: bytes, sheet_name, items, answer_col, detail_col, has_detail) -> bytes:
    wb = load_workbook(io.BytesIO(source), read_only=False, data_only=False)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    review_fill = PatternFill("solid", fgColor="FEF3C7")
    for item in items:
        row = int(item.get("excel_row") or 0)
        if row < 1:
            continue
        if item.get("locked"):
            status_text, detail_text = LOCKED_TEXT, ""
        else:
            status_text, detail_text = _original_cells(item, has_detail)
        c = ws.cell(row=row, column=answer_col, value=status_text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if has_detail and detail_col:
            d = ws.cell(row=row, column=detail_col, value=detail_text)
            d.alignment = Alignment(wrap_text=True, vertical="top")
        if item.get("needs_review"):
            c.fill = review_fill
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fill_csv(source: bytes, items, answer_col, detail_col, has_detail) -> bytes:
    text = source.decode("utf-8-sig", errors="replace")
    rows = [list(r) for r in csv.reader(io.StringIO(text))]
    by_row = {int(it.get("excel_row") or 0): it for it in items}
    a_idx = answer_col - 1
    d_idx = (detail_col - 1) if detail_col else None
    for excel_row, item in by_row.items():
        if excel_row < 1 or excel_row > len(rows):
            continue
        r = rows[excel_row - 1]
        needed = max(a_idx, d_idx if d_idx is not None else a_idx) + 1
        if len(r) < needed:
            r.extend([""] * (needed - len(r)))
        if item.get("locked"):
            status_text, detail_text = LOCKED_TEXT, ""
        else:
            status_text, detail_text = _original_cells(item, has_detail)
        r[a_idx] = status_text
        if d_idx is not None:
            r[d_idx] = detail_text
    out = io.StringIO()
    csv.writer(out).writerows(rows)
    return out.getvalue().encode("utf-8")
