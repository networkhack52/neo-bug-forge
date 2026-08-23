"""Extract the list of questions from an uploaded questionnaire.

Supports .xlsx / .xlsm (openpyxl) and .csv. Vendor security questionnaires
are messy: header rows vary, there are section titles, blank rows, and an
answer column to fill. We auto-detect the question column by scoring each
column on how "question-like" its cells are, then return the cell
coordinates so answers can be written straight back into the same file.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Optional

from openpyxl import load_workbook


@dataclass
class ExtractedQuestion:
    row_index: int          # 0-based logical index across the sheet
    excel_row: int          # 1-based worksheet row (for write-back)
    question: str


@dataclass
class SheetInfo:
    name: str
    question_count: int
    selected: bool = False


@dataclass
class ColumnInfo:
    index: int          # 1-based worksheet column
    header: str         # header cell text (or "Column C" when blank)
    sample: str         # a representative question-like value from the column
    selected: bool = False


@dataclass
class ParseResult:
    questions: list[ExtractedQuestion] = field(default_factory=list)
    question_col: Optional[int] = None   # 1-based worksheet column
    answer_col: Optional[int] = None     # 1-based; status/response goes here
    detail_col: Optional[int] = None     # 1-based; free-text comments go here (if distinct)
    sheet_name: Optional[str] = None
    kind: str = "xlsx"                    # xlsx | csv | text
    # Format-resilience metadata for the confirm-before-quota screen:
    sheets: list = field(default_factory=list)          # SheetInfo per worksheet (xlsx)
    columns: list = field(default_factory=list)         # ColumnInfo candidates in the chosen sheet
    first_questions: list = field(default_factory=list)  # up to 3 extracted questions, for preview
    languages: list = field(default_factory=list)       # detected scripts, e.g. ['en', 'ar']
    rtl: bool = False                                    # any right-to-left script present


_QUESTION_WORDS = (
    "do you", "does", "is ", "are ", "have you", "has ", "can ", "will ", "would ",
    "how ", "what ", "which ", "when ", "where ", "who ",
    "please describe", "please provide", "please confirm", "please specify",
    "describe", "provide", "explain", "list ", "identify", "specify",
    "confirm whether", "indicate whether", "attach",
)

# Verbs that mark a control phrased as a declarative STATEMENT (CAIQ-style items
# to attest to), so "The organization maintains ..." is still captured without a
# question mark — while notes and footers are not.
_CONTROL_VERBS = frozenset((
    "is", "are", "do", "does", "did", "have", "has", "had", "maintain", "maintains",
    "ensure", "ensures", "implement", "implements", "use", "uses", "perform", "performs",
    "provide", "provides", "encrypt", "encrypts", "restrict", "restricts", "review",
    "reviews", "require", "requires", "conduct", "conducts", "support", "supports",
    "store", "stores", "retain", "retains", "monitor", "monitors", "log", "logs",
    "comply", "complies", "follow", "follows", "enforce", "enforces", "document",
    "documents", "protect", "protects", "manage", "manages",
))

# Structure, instructions, and boilerplate — never questions, even mid-sheet.
_NOISE_RE = re.compile(
    r"^\s*(section|part|appendix|domain|category|chapter|table of contents)\b"
    r"|^\s*(instructions?|guidance|overview|introduction)\b"
    r"|^\s*please\s+(complete|answer|fill|read|note|review|see|refer|use|select|"
    r"choose|indicate|rate|mark|ensure)\b"
    r"|^\s*(complete the following|for each|for official use|for internal use|note:)"
    r"|\bpage\s+\d+(\s+of\s+\d+)?\b"
    r"|©|\bcopyright\b|all rights reserved|\bconfidential\b|\bproprietary\b"
    r"|^\s*(version|revision|rev|last updated|last revised)\b[:\s]"
    # Document meta: "Full Question List", "END OF LIST", "(52 questions total)".
    r"|^\s*end of\b|question list\b|\(\d+\s+questions?(\s+total)?\)",
    re.IGNORECASE,
)


def _looks_like_noise(text: str) -> bool:
    """True for section titles, instructions, notes, and footers."""
    t = (text or "").strip()
    if not t or _NOISE_RE.search(t):
        return True
    # A short label/title with no question mark, e.g. "Access Control" or
    # "Encryption:" — a heading, not a question.
    if "?" not in t and len(t.split()) <= 4 and (t.endswith(":") or t.isupper() or t.istitle()):
        return True
    return False


def _has_control_verb(low: str) -> bool:
    return any(tok in _CONTROL_VERBS for tok in re.findall(r"[a-z']+", low))


def _question_score(cell: str) -> float:
    """Higher = more question-like. 0 means 'not a question': headers, section
    titles, instructions, notes, and footers all score 0 so they are never
    counted or answered."""
    if not cell:
        return 0.0
    text = cell.strip()
    if len(text) < 8 or _looks_like_noise(text):
        return 0.0
    low = text.lower()
    words = len(text.split())
    score = 0.0
    if "?" in text:
        score += 2.5
    if any(low.startswith(w) or (" " + w) in low for w in _QUESTION_WORDS):
        score += 1.5
    # No question mark and no lead word: accept only a real declarative control
    # statement (a full sentence with a control verb), so prose notes and footer
    # text do not slip through as questions.
    if score == 0.0:
        if words >= 6 and _has_control_verb(low):
            score += 1.0
        else:
            return 0.0
    if 3 <= words <= 80:
        score += 0.5
    return score


def _looks_like_header(text: str) -> bool:
    t = (text or "").strip().lower().rstrip(":")
    return t in {"question", "questions", "control", "controls", "requirement",
                 "requirements", "item", "items", "description", "control description",
                 "assessment question", "question / requirement"}


# Column a vendor fills with the Yes/No/Partially status.
_ANSWER_HEADERS = {"answer", "response", "vendor response", "compliant", "compliance",
                   "status", "yes/no", "answer (yes/no)"}
# Column a vendor fills with free-text explanation alongside the status.
_DETAIL_HEADERS = {"comments", "comment", "notes", "note", "details", "detail",
                   "additional information", "additional info", "explanation",
                   "description", "vendor comments", "evidence", "remarks"}


def _pick_columns(rows: list[list[str]]) -> tuple[Optional[int], Optional[int], Optional[int]]:
    """Return (question_col, answer_col, detail_col) as 0-based indices.

    ``answer_col`` receives the compliance status; ``detail_col`` (when a
    distinct comments column exists) receives the free-text answer. When there
    is no separate comments column, ``detail_col`` is None and the two are
    written together into ``answer_col``.
    """
    if not rows:
        return None, None, None
    n_cols = max(len(r) for r in rows)
    col_scores = [0.0] * n_cols
    header = rows[0] if rows else []

    for r in rows[:200]:  # sample
        for c in range(n_cols):
            val = r[c] if c < len(r) else ""
            col_scores[c] += _question_score(val)

    # Header hints override weak signals.
    q_col = None
    for c in range(min(n_cols, len(header))):
        if _looks_like_header(header[c]):
            q_col = c
            break
    if q_col is None:
        q_col = max(range(n_cols), key=lambda c: col_scores[c]) if n_cols else None
    if q_col is None or col_scores[q_col] == 0:
        return None, None, None

    def _find_header(names: set[str]) -> Optional[int]:
        for c in range(min(n_cols, len(header))):
            if (header[c] or "").strip().lower() in names:
                return c
        return None

    # Answer (status) column: a matching header, else the column right of Q.
    a_col = _find_header(_ANSWER_HEADERS)
    if a_col is None:
        a_col = q_col + 1

    # Detail (comments) column: only when it's a distinct labelled column.
    d_col = _find_header(_DETAIL_HEADERS)
    if d_col == a_col:
        d_col = None
    return q_col, a_col, d_col


# Right-to-left scripts: Hebrew, Arabic (+ supplements), Syriac, Thaana, and the
# Arabic presentation forms. Used to detect bilingual / RTL questionnaires.
_RTL_RE = re.compile(
    r"[֐-׿؀-ۿ܀-ݏݐ-ݿࢠ-ࣿ"
    r"יִ-﷿ﹰ-﻿]"
)
_ARABIC_RE = re.compile(r"[؀-ۿݐ-ݿࢠ-ࣿﭐ-﷿ﹰ-﻿]")
_HEBREW_RE = re.compile(r"[֐-׿יִ-ﭏ]")
_LATIN_RE = re.compile(r"[A-Za-z]")


def _detect_languages(questions: list[ExtractedQuestion]) -> tuple[list[str], bool]:
    """Best-effort script detection across the extracted questions. Returns
    (languages, rtl). We answer in English regardless; this only NOTES what the
    source used so a bilingual sheet is handled transparently (no Arabic drafting)."""
    text = "\n".join(q.question for q in questions)
    langs: list[str] = []
    if _LATIN_RE.search(text):
        langs.append("en")
    if _ARABIC_RE.search(text):
        langs.append("ar")
    if _HEBREW_RE.search(text):
        langs.append("he")
    return langs, bool(_RTL_RE.search(text))


def _extract_questions(rows: list[list[str]], q_col: int) -> list[ExtractedQuestion]:
    """Pull the real questions out of one column, skipping the header row, section
    titles, instructions, notes, and footers (they score 0)."""
    out: list[ExtractedQuestion] = []
    logical = 0
    header_seen = False
    for excel_row, r in enumerate(rows, start=1):
        cell = r[q_col] if q_col < len(r) else ""
        if _question_score(cell) <= 0:
            continue
        if not header_seen and _looks_like_header(cell):
            header_seen = True
            continue
        out.append(ExtractedQuestion(row_index=logical, excel_row=excel_row, question=cell.strip()))
        logical += 1
    return out


def _column_candidates(rows: list[list[str]], chosen: Optional[int]) -> list[ColumnInfo]:
    """Every column that holds at least one question-like cell, so the user can
    override the auto-detected question column on the confirm screen."""
    if not rows:
        return []
    n_cols = max(len(r) for r in rows)
    header = rows[0] if rows else []
    out: list[ColumnInfo] = []
    for c in range(n_cols):
        best = ""
        best_score = 0.0
        for r in rows[:200]:
            val = r[c] if c < len(r) else ""
            s = _question_score(val)
            if s > best_score:
                best_score, best = s, val.strip()
        if best_score <= 0:
            continue
        head = (header[c].strip() if c < len(header) and header[c] else "") or f"Column {chr(65 + c) if c < 26 else c + 1}"
        out.append(ColumnInfo(index=c + 1, header=head,
                              sample=best[:120], selected=(chosen is not None and c == chosen)))
    return out


def _finish_result(result: ParseResult, rows: list[list[str]], q_col: Optional[int]) -> ParseResult:
    """Attach the confirm-screen metadata (column candidates, preview, languages)."""
    result.columns = _column_candidates(rows, q_col)
    result.first_questions = [q.question for q in result.questions[:3]]
    result.languages, result.rtl = _detect_languages(result.questions)
    return result


def parse_xlsx(data: bytes, sheet: Optional[str] = None,
               force_col: Optional[int] = None) -> ParseResult:
    """Parse an .xlsx/.xlsm. Scores EVERY worksheet and answers the one with the
    most real questions (cover sheets and instruction tabs score ~0), listing the
    others so the user can switch. ``sheet``/``force_col`` (1-based) override the
    auto-detected worksheet and question column from the confirm screen."""
    wb = load_workbook(io.BytesIO(data), read_only=False, data_only=True)

    # Read each worksheet's rows once and score it.
    per_sheet: dict[str, list[list[str]]] = {}
    scored: list[tuple[str, int, Optional[int], Optional[int], Optional[int]]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [["" if v is None else str(v) for v in row] for row in ws.iter_rows(values_only=True)]
        per_sheet[name] = rows
        qc, ac, dc = _pick_columns(rows)
        n = len(_extract_questions(rows, qc)) if qc is not None else 0
        scored.append((name, n, qc, ac, dc))

    if not scored:
        return ParseResult(kind="xlsx")

    # Choose the requested sheet, else the one with the most questions.
    chosen = None
    if sheet:
        chosen = next((s for s in scored if s[0] == sheet), None)
    if chosen is None:
        chosen = max(scored, key=lambda s: s[1])
    name, _n, q_col, a_col, d_col = chosen
    if force_col is not None:
        q_col = force_col - 1  # 1-based -> 0-based
    rows = per_sheet[name]

    result = ParseResult(
        question_col=None if q_col is None else q_col + 1,
        answer_col=None if a_col is None else a_col + 1,
        detail_col=None if d_col is None else d_col + 1,
        sheet_name=name, kind="xlsx",
        sheets=[SheetInfo(name=s[0], question_count=s[1], selected=(s[0] == name)) for s in scored],
    )
    if q_col is None:
        return _finish_result(result, rows, None)
    result.questions = _extract_questions(rows, q_col)
    return _finish_result(result, rows, q_col)


def parse_csv(data: bytes, force_col: Optional[int] = None) -> ParseResult:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [[c for c in row] for row in reader]
    q_col, a_col, d_col = _pick_columns(rows)
    if force_col is not None:
        q_col = force_col - 1  # 1-based -> 0-based override from the confirm screen
    result = ParseResult(question_col=None if q_col is None else q_col + 1,
                         answer_col=None if a_col is None else a_col + 1,
                         detail_col=None if d_col is None else d_col + 1,
                         kind="csv")
    if q_col is None:
        return _finish_result(result, rows, None)
    result.questions = _extract_questions(rows, q_col)
    return _finish_result(result, rows, q_col)


# Leading list markers on a plain-text question line: "1. ", "12) ", "- ", "• ".
_LIST_PREFIX = re.compile(r"^\s*(?:\d{1,3}[.)]|[-*•·])\s+")


def parse_text(data: bytes) -> ParseResult:
    """Plain-text questionnaire: one question per line. Unlike CSV, this never
    splits on the commas *inside* a question, so questions stay whole. Strips
    leading list numbering and skips section headers, separators, and notes."""
    text = data.decode("utf-8-sig", errors="replace")
    result = ParseResult(kind="text")
    logical = 0
    for line_no, raw in enumerate(text.splitlines(), start=1):
        line = _LIST_PREFIX.sub("", raw.strip()).strip()
        if _question_score(line) <= 0:
            continue
        result.questions.append(
            ExtractedQuestion(row_index=logical, excel_row=line_no, question=line)
        )
        logical += 1
    # No columns in plain text, but still surface the preview + languages.
    result.first_questions = [q.question for q in result.questions[:3]]
    result.languages, result.rtl = _detect_languages(result.questions)
    return result


def parse(filename: str, data: bytes, sheet: Optional[str] = None,
          force_col: Optional[int] = None) -> ParseResult:
    """Route by extension. ``sheet`` and ``force_col`` (1-based) let the confirm
    screen re-parse with a user-chosen worksheet / question column."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return parse_csv(data, force_col=force_col)
    if name.endswith((".xlsx", ".xlsm")):
        return parse_xlsx(data, sheet=sheet, force_col=force_col)
    if name.endswith((".txt", ".md", ".text")):
        return parse_text(data)
    # Unknown extension: try xlsx; else pick text vs csv by how comma-delimited it
    # looks, so a plain-text list isn't mangled by comma-splitting.
    try:
        return parse_xlsx(data, sheet=sheet, force_col=force_col)
    except Exception:
        pass
    try:
        sample = data.decode("utf-8-sig", errors="replace")[:5000]
        lines = [ln for ln in sample.splitlines() if ln.strip()]
        if lines and sum(1 for ln in lines if "," in ln) / len(lines) >= 0.5:
            return parse_csv(data, force_col=force_col)
    except Exception:
        pass
    return parse_text(data)
