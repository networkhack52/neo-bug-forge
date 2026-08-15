import io

from openpyxl import Workbook, load_workbook

from app import export, parsing


def _wb_bytes(rows):
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_detects_distinct_detail_column():
    rows = [
        ["Question", "Response", "Comments"],
        ["Do you enforce MFA for all employees?", "", ""],
    ]
    result = parsing.parse_xlsx(_wb_bytes(rows))
    assert result.question_col == 1
    assert result.answer_col == 2   # status column
    assert result.detail_col == 3   # free-text column


def test_export_original_fills_status_and_detail_columns():
    rows = [
        ["Question", "Response", "Comments"],
        ["Do you enforce MFA for all employees?", "", ""],
        ["Is customer data encrypted at rest?", "", ""],
    ]
    source = _wb_bytes(rows)
    parsed = parsing.parse_xlsx(source)

    q = {
        "source_bytes": source,
        "source_kind": "xlsx",
        "sheet_name": parsed.sheet_name,
        "answer_col": parsed.answer_col,
        "detail_col": parsed.detail_col,
    }
    items = [
        {"excel_row": parsed.questions[0].excel_row, "choice": "Yes",
         "answer": "MFA is enforced via Okta.", "needs_review": False},
        {"excel_row": parsed.questions[1].excel_row, "choice": "Yes",
         "answer": "AES-256 at rest.", "needs_review": True},
    ]
    out = export.export_original(q, items)
    ws = load_workbook(io.BytesIO(out)).active
    assert ws.cell(row=parsed.questions[0].excel_row, column=2).value == "Yes"
    assert ws.cell(row=parsed.questions[0].excel_row, column=3).value == "MFA is enforced via Okta."
    assert ws.cell(row=parsed.questions[1].excel_row, column=2).value == "Yes"


def test_export_original_combines_when_no_detail_column():
    rows = [
        ["Question", "Answer"],
        ["Do you enforce MFA?", ""],
    ]
    source = _wb_bytes(rows)
    parsed = parsing.parse_xlsx(source)
    assert parsed.detail_col is None
    q = {
        "source_bytes": source, "source_kind": "xlsx", "sheet_name": parsed.sheet_name,
        "answer_col": parsed.answer_col, "detail_col": parsed.detail_col,
    }
    items = [{"excel_row": parsed.questions[0].excel_row, "choice": "Yes",
              "answer": "Enforced via Okta.", "needs_review": False}]
    ws = load_workbook(io.BytesIO(export.export_original(q, items))).active
    assert ws.cell(row=parsed.questions[0].excel_row, column=2).value == "Yes. Enforced via Okta."


def test_export_original_csv_roundtrip():
    csv_bytes = (
        "Question,Response,Comments\n"
        "Do you encrypt data in transit?,,\n"
    ).encode()
    parsed = parsing.parse_csv(csv_bytes)
    q = {
        "source_bytes": csv_bytes, "source_kind": "csv", "sheet_name": None,
        "answer_col": parsed.answer_col, "detail_col": parsed.detail_col,
    }
    items = [{"excel_row": parsed.questions[0].excel_row, "choice": "Yes",
              "answer": "TLS 1.2+.", "needs_review": False}]
    out = export.export_original(q, items).decode()
    assert "Yes" in out
    assert "TLS 1.2+." in out


def test_clean_answer_strips_internal_notes():
    dirty = ("We store data in the EU.\n\n[Attestly review: these claims were not fully "
             "supported by your Answer Library — verify before sending: eligible customers]")
    assert export.clean_answer(dirty) == "We store data in the EU."
    assert export.clean_answer("Yes.\n\n[Attestly: model call failed: TimeoutException]") == "Yes."


def test_export_simple_carries_source_and_status_columns():
    items = [
        {"question": "Is data encrypted at rest?", "choice": "Yes",
         "answer": "Yes. AES-256 everywhere.", "needs_review": False, "locked": 0,
         "citations": '[{"title": "SOC2.pdf", "text": "encrypted at rest with AES-256", "kind": "document"}]'},
        {"question": "Do you run chaos drills?", "choice": "",
         "answer": export.NO_EVIDENCE if hasattr(export, "NO_EVIDENCE") else "No supporting evidence.",
         "needs_review": True, "locked": 0, "citations": "[]"},
        {"question": "Control 999?", "choice": "", "answer": "", "needs_review": True, "locked": 1,
         "citations": "[]"},
    ]
    ws = load_workbook(io.BytesIO(export.export_simple("q", items))).active
    assert [ws.cell(row=1, column=c).value for c in range(1, 6)] == \
        ["Section", "Question", "Vendor Response", "Source", "Status"]
    # Answered row: has a Source and Status Answered.
    assert "SOC2.pdf" in ws.cell(row=2, column=4).value
    assert ws.cell(row=2, column=5).value == "Answered"
    # Ungrounded row: no source, Status No evidence.
    assert ws.cell(row=3, column=5).value == "No evidence"
    # Locked row: marked Locked, response says so.
    assert ws.cell(row=4, column=5).value == "Locked"
    assert "Locked" in ws.cell(row=4, column=3).value


def test_cell_text_avoids_duplicate_status_prefix():
    # Answer already leads with the status -> no "Yes — Yes." doubling.
    assert export._cell_text("Yes", "Yes. MFA is enforced.", False) == ("Yes. MFA is enforced.", "")
    # Answer doesn't lead with the status -> prepend it.
    assert export._cell_text("Yes", "MFA is enforced.", False) == ("Yes. MFA is enforced.", "")
    # Internal notes are stripped here too.
    assert export._cell_text("No", "No. [Attestly review: check]", False)[0] == "No."


def test_can_export_original_flag():
    assert export.can_export_original({"source_bytes": b"x", "answer_col": 2}) is True
    assert export.can_export_original({"source_bytes": None, "answer_col": 2}) is False
    assert export.can_export_original({"source_bytes": b"x", "answer_col": None}) is False
