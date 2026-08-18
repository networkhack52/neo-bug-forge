import io
import uuid

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app

client = TestClient(app)


def _token(domain=None):
    # Unique email per call. Default to a unique DOMAIN too, so each org gets a
    # fresh onboarding allowance (the 150 pool is shared per email domain).
    domain = domain or f"{uuid.uuid4().hex[:12]}.example.com"
    email = f"t-{uuid.uuid4().hex[:8]}@{domain}"
    r = client.post("/v1/signup", json={"name": "Test Co", "email": email, "password": "supersecret"})
    assert r.status_code == 200, r.text
    return r.json()["api_token"]


def _auth(token):
    return {"Authorization": f"Bearer {token}"}


def _questionnaire_bytes():
    wb = Workbook()
    ws = wb.active
    ws.append(["Question", "Answer"])
    ws.append(["Do you enforce MFA for all employees?", ""])
    ws.append(["Is data encrypted in transit?", ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload_doc(token, text="Security policy: access is least privilege; data encrypted at rest with AES-256."):
    files = {"file": ("policy.txt", text.encode(), "text/plain")}
    r = client.post("/v1/documents", headers=_auth(token), files=files)
    assert r.status_code == 200, r.text


def test_health():
    r = client.get("/health")
    assert r.status_code == 200
    assert r.json()["status"] == "ok"


def test_seed_starter_loads_and_is_idempotent():
    token = _token()
    r1 = client.post("/v1/answers/seed_starter", headers=_auth(token)).json()
    assert r1["created"] > 20            # the bundled starter set
    assert r1["bank_size"] == r1["created"]
    # Second call adds nothing (everything already present / deduped).
    r2 = client.post("/v1/answers/seed_starter", headers=_auth(token)).json()
    assert r2["created"] == 0
    assert r2["skipped"] >= r1["created"]


def test_oversized_upload_is_rejected(monkeypatch):
    from app import config

    monkeypatch.setattr(config, "MAX_UPLOAD_BYTES", 1024)  # 1 KB cap for the test
    token = _token()
    big = b"x" * 2048
    r = client.post("/v1/documents", headers=_auth(token),
                    files={"file": ("big.txt", big, "text/plain")})
    assert r.status_code == 413


def test_auth_required():
    assert client.get("/v1/me").status_code == 401
    assert client.get("/v1/me", headers=_auth("bogus")).status_code == 401


def test_cors_locked_to_known_origins():
    ok = client.get("/health", headers={"Origin": "https://attestly-gamma.vercel.app"})
    assert ok.headers.get("access-control-allow-origin") == "https://attestly-gamma.vercel.app"
    # An unknown origin is not echoed back (no wildcard).
    bad = client.get("/health", headers={"Origin": "https://evil.example.com"})
    assert bad.headers.get("access-control-allow-origin") not in ("*", "https://evil.example.com")


def test_security_headers_present():
    h = client.get("/health").headers
    assert h.get("x-content-type-options") == "nosniff"
    assert h.get("x-frame-options") == "DENY"


def test_signup_requires_email_and_password():
    assert client.post("/v1/signup", json={"name": "X", "email": "bad", "password": "supersecret"}).status_code == 400
    assert client.post("/v1/signup", json={"name": "X", "email": "a@b.com", "password": "short"}).status_code == 400


def test_signup_dedupes_email_and_login_returns_same_org():
    email = f"owner-{uuid.uuid4().hex[:8]}@example.com"
    r1 = client.post("/v1/signup", json={"name": "Acme", "email": email, "password": "supersecret"})
    assert r1.status_code == 200
    org_id = r1.json()["org_id"]
    # Second signup with the same email is rejected.
    assert client.post("/v1/signup", json={"name": "Acme2", "email": email, "password": "supersecret"}).status_code == 409
    # Login with the right password returns the SAME org (and its data/token).
    ok = client.post("/v1/login", json={"email": email, "password": "supersecret"})
    assert ok.status_code == 200
    assert ok.json()["org_id"] == org_id
    # Wrong password is rejected.
    assert client.post("/v1/login", json={"email": email, "password": "nope"}).status_code == 401
    # Unknown email is rejected the same way.
    assert client.post("/v1/login", json={"email": "ghost@example.com", "password": "x"}).status_code == 401


def test_full_flow_signup_bank_upload_export():
    token = _token()

    # New org starts on free tier with empty bank.
    me = client.get("/v1/me", headers=_auth(token)).json()
    assert me["tier"] == "free"
    assert me["bank_size"] == 0

    # Seed the bank with the two answers the questionnaire will ask.
    client.post("/v1/answers", headers=_auth(token), json={
        "question": "Do you enforce multi-factor authentication (MFA) for all employees?",
        "answer": "Yes, MFA is enforced for all employees.",
    })
    client.post("/v1/answers", headers=_auth(token), json={
        "question": "Is data encrypted in transit?",
        "answer": "Yes, TLS 1.2+ everywhere.",
    })

    # Upload a questionnaire -> auto-answered.
    files = {"file": ("q.xlsx", _questionnaire_bytes(),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = client.post("/v1/questionnaires", headers=_auth(token), files=files)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["total_questions"] == 2
    assert body["reused_from_bank"] >= 1   # at least the verbatim MFA question

    qid = body["questionnaire_id"]
    assert body["answered"] == 2
    assert body["locked"] == 0

    # Metering drew from the one-time onboarding pool first (period untouched).
    me = client.get("/v1/me", headers=_auth(token)).json()
    assert me["questions_used"] == 0
    assert me["onboarding_remaining"] == 148          # 150 - 2
    assert me["answers_remaining"] == 173             # 148 + 25

    # Export returns a real xlsx.
    r = client.get(f"/v1/questionnaires/{qid}/export", headers=_auth(token))
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert len(r.content) > 0


def _big_questionnaire(n):
    wb = Workbook()
    ws = wb.active
    ws.append(["Question", "Answer"])
    for i in range(n):
        ws.append([f"Do you support control number {i}?", ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_onboarding_allowance_answers_beyond_the_free_period_limit():
    # A brand-new free account (with a trust doc) can run a big questionnaire
    # (>25) in one go, because the 150 onboarding pool covers it — nothing declined.
    token = _token()
    _upload_doc(token)  # unlocks drafting
    files = {"file": ("big.xlsx", _big_questionnaire(30),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = client.post("/v1/questionnaires", headers=_auth(token), files=files)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["answered"] == 30 and body["locked"] == 0
    me = client.get("/v1/me", headers=_auth(token)).json()
    assert me["questions_used"] == 0            # drawn from onboarding, not the period
    assert me["onboarding_remaining"] == 120    # 150 - 30


def test_over_quota_partial_answers_and_locks_the_rest(monkeypatch):
    from app import config
    monkeypatch.setattr(config, "ONBOARDING_ALLOWANCE", 0)  # only the 25 period remains
    token = _token()
    _upload_doc(token)
    files = {"file": ("big.xlsx", _big_questionnaire(30),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = client.post("/v1/questionnaires", headers=_auth(token), files=files)
    assert r.status_code == 200, r.text          # partial, never declined
    body = r.json()
    assert body["answered"] == 25 and body["locked"] == 5
    # The export is still produced, with the locked rows in it.
    ex = client.get(f"/v1/questionnaires/{body['questionnaire_id']}/export", headers=_auth(token))
    assert ex.status_code == 200 and len(ex.content) > 0


def test_onboarding_allowance_is_shared_per_domain(monkeypatch):
    from app import config
    monkeypatch.setattr(config, "ONBOARDING_ALLOWANCE", 40)
    domain = f"{uuid.uuid4().hex[:12]}.acme.test"
    a = _token(domain=domain)
    b = _token(domain=domain)  # same company, second signup
    _upload_doc(a)
    files = {"file": ("big.xlsx", _big_questionnaire(30),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    client.post("/v1/questionnaires", headers=_auth(a), files=files)
    # Account A spent 30 of the shared 40; account B sees only 10 left in the pool.
    me_b = client.get("/v1/me", headers=_auth(b)).json()
    assert me_b["onboarding_remaining"] == 10


def test_usage_cost_endpoint_starts_at_zero():
    token = _token()
    c = client.get("/v1/usage/cost", headers=_auth(token)).json()
    assert c["answers"] == 0 and c["cost_usd"] == 0


def _org_id(token):
    return client.get("/v1/me", headers=_auth(token)).json()["org_id"]


def test_signup_always_rejects_disposable_domains():
    r = client.post("/v1/signup", json={"name": "X", "email": "a@mailinator.com", "password": "supersecret"})
    assert r.status_code == 400


def test_signup_allows_gmail_by_default_but_blocks_it_when_work_email_required(monkeypatch):
    from app import config
    # Default: Gmail is allowed so founders/evaluators can sign up.
    monkeypatch.setattr(config, "REQUIRE_WORK_EMAIL", False)
    ok = client.post("/v1/signup", json={"name": "X", "email": f"a-{uuid.uuid4().hex[:6]}@gmail.com", "password": "supersecret"})
    assert ok.status_code == 200
    # Strict mode on: Gmail is rejected.
    monkeypatch.setattr(config, "REQUIRE_WORK_EMAIL", True)
    blocked = client.post("/v1/signup", json={"name": "X", "email": "b@gmail.com", "password": "supersecret"})
    assert blocked.status_code == 400


def test_free_tier_spend_cap_pauses_drafting(monkeypatch):
    from app import config, db
    monkeypatch.setattr(config, "FREE_TIER_MONTHLY_SPEND_CAP_USD", 0.01)
    token = _token()
    _upload_doc(token)  # would normally unlock drafting
    # Simulate free-tier spend already over the cap.
    db.record_usage(_org_id(token), None, None,
                    {"model": "claude-haiku-4-5-20251001", "input_tokens": 0,
                     "cached_input_tokens": 0, "output_tokens": 0}, cost=1.0)
    files = {"file": ("q.xlsx", _big_questionnaire(3),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    body = client.post("/v1/questionnaires", headers=_auth(token), files=files).json()
    assert body["spend_paused"] is True
    assert body["answered"] == 0 and body["blocked"] == 3


def test_upload_is_rate_limited(monkeypatch):
    from app import config, ratelimit
    monkeypatch.setattr(config, "RATE_LIMIT_ENABLED", True)
    monkeypatch.setattr(config, "RL_UPLOAD", (3, 3600))
    ratelimit.reset()
    token = _token()

    def one():
        files = {"file": ("q.xlsx", _big_questionnaire(1),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        return client.post("/v1/questionnaires", headers=_auth(token), files=files).status_code

    codes = [one() for _ in range(4)]
    assert codes[:3] == [200, 200, 200]
    assert codes[3] == 429
    ratelimit.reset()


def test_free_tier_gates_drafting_until_a_document_is_uploaded():
    # No document -> questions that need drafting are blocked and cost nothing.
    token = _token()
    files = {"file": ("q.xlsx", _big_questionnaire(3),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    body = client.post("/v1/questionnaires", headers=_auth(token), files=files).json()
    assert body["answered"] == 0 and body["blocked"] == 3
    assert client.get("/v1/me", headers=_auth(token)).json()["onboarding_remaining"] == 150
    # After a document, the same questions draft normally.
    _upload_doc(token)
    body2 = client.post("/v1/questionnaires", headers=_auth(token), files=files).json()
    assert body2["answered"] == 3 and body2["blocked"] == 0


def test_simulated_upgrade_changes_tier():
    token = _token()
    r = client.post("/v1/billing/checkout", headers=_auth(token), json={"tier": "starter"})
    assert r.status_code == 200
    assert r.json()["simulated"] is True
    r = client.post("/v1/billing/confirm", headers=_auth(token), json={"tier": "starter"})
    assert r.status_code == 200
    assert r.json()["tier"] == "starter"
    assert r.json()["question_limit"] == 750


def test_annual_checkout_charges_yearly_price():
    token = _token()
    r = client.post("/v1/billing/checkout", headers=_auth(token),
                    json={"tier": "growth", "interval": "year"})
    assert r.status_code == 200
    body = r.json()
    assert body["interval"] == "year"
    assert body["amount"] == 2490  # 10 x $249, 2 months free


def test_invalid_interval_rejected():
    token = _token()
    r = client.post("/v1/billing/checkout", headers=_auth(token),
                    json={"tier": "starter", "interval": "weekly"})
    assert r.status_code == 400


def test_resolve_contradiction_keep_new_updates_library():
    from app import db

    token = _token()
    q = "How often do you rotate encryption keys?"
    ans = client.post("/v1/answers", headers=_auth(token),
                      json={"question": q, "answer": "We rotate keys annually."}).json()
    files = {"file": ("q.xlsx", _one_q_bytes(q),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    body = client.post("/v1/questionnaires", headers=_auth(token), files=files).json()
    item_id = body["items"][0]["id"]

    # Simulate a flagged contradiction on the item (new drafted answer differs).
    db.set_item_answer_text(item_id, "We rotate keys quarterly.", "")
    db.set_item_contradiction(item_id, {
        "answer_id": ans["id"], "question": q,
        "prior_answer": "We rotate keys annually.",
        "new_answer": "We rotate keys quarterly.",
        "reason": "frequency changed",
    })

    r = client.post(f"/v1/items/{item_id}/resolve_contradiction",
                    headers=_auth(token), json={"keep": "new"})
    assert r.status_code == 200
    # The library entry now reflects the new answer, with a change log.
    updated = db.get_answer(ans["id"])
    assert "quarterly" in updated["answer"]
    import json as _json
    assert len(_json.loads(updated["change_log"])) == 1
    # Flag is cleared.
    assert (db.get_item(item_id)["contradiction"] or "") == ""


def test_resolve_contradiction_keep_prior_reverts_item():
    from app import db

    token = _token()
    q = "What is your data retention period?"
    ans = client.post("/v1/answers", headers=_auth(token),
                      json={"question": q, "answer": "We retain data for 30 days."}).json()
    files = {"file": ("q.xlsx", _one_q_bytes(q),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    body = client.post("/v1/questionnaires", headers=_auth(token), files=files).json()
    item_id = body["items"][0]["id"]

    db.set_item_answer_text(item_id, "We retain data for 90 days.", "")
    db.set_item_contradiction(item_id, {
        "answer_id": ans["id"], "question": q,
        "prior_answer": "We retain data for 30 days.",
        "new_answer": "We retain data for 90 days.",
        "reason": "a number or time period changed",
    })

    r = client.post(f"/v1/items/{item_id}/resolve_contradiction",
                    headers=_auth(token), json={"keep": "prior"})
    assert r.status_code == 200
    # Item reverted to the prior library answer; library unchanged.
    assert "30 days" in db.get_item(item_id)["answer"]
    assert "30 days" in db.get_answer(ans["id"])["answer"]
    assert (db.get_item(item_id)["contradiction"] or "") == ""


def _one_q_bytes(question):
    wb = Workbook()
    ws = wb.active
    ws.append(["Question", "Answer"])
    ws.append([question, ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_export_blocks_bare_no_until_remediation_is_set():
    token = _token()
    q = "Do you sell customer data to third parties?"
    # A reused library answer that reads as a 'No' -> the item's choice is No.
    client.post("/v1/answers", headers=_auth(token),
                json={"question": q, "answer": "No, we do not sell customer data."})
    files = {"file": ("q.xlsx", _one_q_bytes(q),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    body = client.post("/v1/questionnaires", headers=_auth(token), files=files).json()
    qid = body["questionnaire_id"]
    item = body["items"][0]
    assert item["choice"] == "No"

    # Export is blocked while the 'No' is bare.
    r = client.get(f"/v1/questionnaires/{qid}/export", headers=_auth(token))
    assert r.status_code == 409
    assert r.json()["unresolved"][0]["issue"] == "remediation"

    # Add a remediation date -> export succeeds.
    client.post(f"/v1/items/{item['id']}/remediation", headers=_auth(token),
                json={"remediation_date": "2026-03-01"})
    r = client.get(f"/v1/questionnaires/{qid}/export", headers=_auth(token))
    assert r.status_code == 200

    # A no-plan acknowledgement also clears the gate.
    client.post(f"/v1/items/{item['id']}/remediation", headers=_auth(token),
                json={"no_plan": True})
    assert client.get(f"/v1/questionnaires/{qid}/export", headers=_auth(token)).status_code == 200


# --- Password reset -------------------------------------------------------
def _capture_reset_link(monkeypatch):
    """Intercept the outbound email so tests can read the reset link."""
    captured = {}

    def fake_send(to, reset_url):
        captured["to"] = to
        captured["url"] = reset_url
        return True

    monkeypatch.setattr("app.main.email_svc.send_password_reset", fake_send)
    return captured


def test_password_reset_full_flow(monkeypatch):
    email = f"reset-{uuid.uuid4().hex[:8]}@example.com"
    r = client.post("/v1/signup", json={"name": "Reset Co", "email": email, "password": "originalpw"})
    assert r.status_code == 200
    old_token = r.json()["api_token"]

    captured = _capture_reset_link(monkeypatch)
    r = client.post("/v1/password/forgot", json={"email": email})
    assert r.status_code == 200
    assert captured["to"] == email
    reset_token = captured["url"].split("reset=")[1]

    # Reset to a new password; returns a fresh usable token.
    r = client.post("/v1/password/reset", json={"token": reset_token, "password": "brandnewpw"})
    assert r.status_code == 200, r.text
    new_token = r.json()["api_token"]
    assert client.get("/v1/me", headers=_auth(new_token)).status_code == 200
    # Old API token was rotated out.
    assert client.get("/v1/me", headers=_auth(old_token)).status_code == 401
    # Old password no longer logs in; new one does.
    assert client.post("/v1/login", json={"email": email, "password": "originalpw"}).status_code == 401
    assert client.post("/v1/login", json={"email": email, "password": "brandnewpw"}).status_code == 200


def test_password_reset_token_is_single_use(monkeypatch):
    email = f"once-{uuid.uuid4().hex[:8]}@example.com"
    client.post("/v1/signup", json={"name": "Once Co", "email": email, "password": "originalpw"})
    captured = _capture_reset_link(monkeypatch)
    client.post("/v1/password/forgot", json={"email": email})
    token = captured["url"].split("reset=")[1]
    assert client.post("/v1/password/reset", json={"token": token, "password": "firstresetpw"}).status_code == 200
    # A second use of the same link is rejected.
    assert client.post("/v1/password/reset", json={"token": token, "password": "secondtry"}).status_code == 400


def test_password_forgot_does_not_enumerate_accounts(monkeypatch):
    _capture_reset_link(monkeypatch)
    # Unknown email returns the same 200 + generic message as a known one.
    r = client.post("/v1/password/forgot", json={"email": f"nobody-{uuid.uuid4().hex}@example.com"})
    assert r.status_code == 200
    assert "reset link" in r.json()["message"].lower()


def test_password_reset_rejects_bad_token():
    r = client.post("/v1/password/reset", json={"token": "atl_not_a_real_token", "password": "longenough"})
    assert r.status_code == 400


def test_password_reset_requires_min_length(monkeypatch):
    email = f"short-{uuid.uuid4().hex[:8]}@example.com"
    client.post("/v1/signup", json={"name": "Short Co", "email": email, "password": "originalpw"})
    captured = _capture_reset_link(monkeypatch)
    client.post("/v1/password/forgot", json={"email": email})
    token = captured["url"].split("reset=")[1]
    assert client.post("/v1/password/reset", json={"token": token, "password": "short"}).status_code == 400
