import io

from openpyxl import Workbook

from app import parsing


def _wb_bytes(rows):
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_only_real_questions_are_extracted_from_a_messy_sheet():
    rows = [
        ["Question", "Answer"],
        ["Vendor Security Assessment 2025", ""],                 # title
        ["1. Access Control", ""],                               # section header
        ["Please answer all questions in this section.", ""],    # instruction
        ["Do you enforce MFA for all employees?", ""],           # REAL question
        ["The organization maintains a documented access control policy.", ""],  # control statement
        ["2. Encryption", ""],                                   # section header
        ["Is data encrypted at rest?", ""],                      # REAL question
        ["Confidential - Acme Corp - Page 1 of 3", ""],          # footer
        ["© 2025 Acme Corp. All rights reserved.", ""],          # boilerplate
    ]
    result = parsing.parse_xlsx(_wb_bytes(rows))
    got = [q.question for q in result.questions]
    assert got == [
        "Do you enforce MFA for all employees?",
        "The organization maintains a documented access control policy.",
        "Is data encrypted at rest?",
    ], got


def test_section_titles_and_footers_score_zero():
    for noise in ["Access Control", "ENCRYPTION", "Section 3: Governance",
                  "Page 2 of 10", "© 2025 Acme", "Instructions", "Overview",
                  "Please complete the following table.", "For each control, provide evidence.",
                  "This document is confidential and proprietary."]:
        assert parsing._question_score(noise) == 0.0, noise


def test_real_questions_and_control_statements_score_positive():
    for q in ["Do you enforce MFA for all employees?",
              "Is data encrypted at rest?",
              "Describe your incident response process.",
              "The organization maintains a documented data classification policy.",
              "We encrypt all customer data at rest and in transit."]:
        assert parsing._question_score(q) > 0.0, q


def test_bundled_sample_questionnaire_parses_to_eight_questions():
    # The onboarding "Try a sample" file must parse cleanly to its 8 questions.
    import pathlib

    sample = pathlib.Path(__file__).resolve().parents[2] / "frontend" / "public" / "sample-questionnaire.xlsx"
    assert sample.exists(), sample
    result = parsing.parse_xlsx(sample.read_bytes())
    assert len(result.questions) == 8
    assert all("?" in q.question or q.question.lower().startswith(("how", "do", "is", "are"))
               for q in result.questions)


def test_plain_text_questionnaire_keeps_questions_whole():
    # A .txt list: numbered lines, commas INSIDE questions, section headers,
    # separators, and metadata. Must extract full questions, not comma-truncate.
    txt = (
        "OSSTMM 4 Vendor Security Checklist\n"
        "Full Question List (3 questions)\n"
        "========================================\n"
        "SECTION 1 - TRUST (2 questions)\n"
        "1. Does the contract define mutual obligations, or does the vendor accept terms one-way?\n"
        "2. Can you observe the vendor's operations relevant to your data (audit reports, logs, dashboards)?\n"
        "SECTION 2 - OPERATIONS\n"
        "3. Does the vendor require phishing-resistant MFA for all administrative access?\n"
        "END OF LIST (3 questions total)\n"
    )
    result = parsing.parse("checklist.txt", txt.encode())
    assert result.kind == "text"
    got = [q.question for q in result.questions]
    assert got == [
        "Does the contract define mutual obligations, or does the vendor accept terms one-way?",
        "Can you observe the vendor's operations relevant to your data (audit reports, logs, dashboards)?",
        "Does the vendor require phishing-resistant MFA for all administrative access?",
    ], got


def _multi_sheet_bytes():
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    for r in [["Vendor Security Questionnaire"], ["Prepared for Acme Corp"], ["Confidential"]]:
        cover.append(r)
    ws = wb.create_sheet("Controls")
    for r in [["#", "Question", "Response"],
              ["1", "Do you encrypt data at rest?", ""],
              ["2", "Is MFA enforced for all staff?", ""],
              ["3", "Do you run annual penetration tests?", ""]]:
        ws.append(r)
    notes = wb.create_sheet("Instructions")
    notes.append(["Please complete the Controls tab and return it."])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_multi_sheet_picks_the_controls_tab_and_lists_the_others():
    result = parsing.parse_xlsx(_multi_sheet_bytes())
    assert result.sheet_name == "Controls"
    assert len(result.questions) == 3
    names = {s.name for s in result.sheets}
    assert names == {"Cover", "Controls", "Instructions"}
    selected = [s for s in result.sheets if s.selected]
    assert len(selected) == 1 and selected[0].name == "Controls"
    # The confirm screen gets a 3-question preview + the chosen 1-based column.
    assert result.first_questions[0] == "Do you encrypt data at rest?"
    assert result.question_col == 2  # 'Question' is the 2nd column


def test_column_override_reparses_on_the_chosen_column():
    data = _multi_sheet_bytes()
    # Force column 1 ('#') — it holds no questions, so extraction yields nothing.
    forced = parsing.parse_xlsx(data, sheet="Controls", force_col=1)
    assert forced.question_col == 1
    assert len(forced.questions) == 0
    # The candidate list still offers the real question column to switch back to.
    assert any(c.index == 2 for c in forced.columns)


def test_rtl_and_bilingual_detection():
    wb = Workbook()
    ws = wb.active
    ws.append(["Question", "Answer"])
    ws.append(["Do you encrypt data at rest? هل تقومون بتشفير البيانات؟", ""])
    ws.append(["Is MFA enforced? هل يتم فرض المصادقة الثنائية؟", ""])
    buf = io.BytesIO()
    wb.save(buf)
    result = parsing.parse_xlsx(buf.getvalue())
    assert result.rtl is True
    assert "ar" in result.languages and "en" in result.languages


def test_csv_messy_sheet_extracts_only_questions():
    csv_text = (
        "Question,Answer\n"
        "Information Security Questionnaire,\n"          # title
        "Domain: Access Management,\n"                   # section
        "Do you review user access periodically?,\n"     # REAL
        "Please see the guidance notes above.,\n"        # instruction
        "Are audit logs retained?,\n"                    # REAL
        "Page 1,\n"                                      # footer
    )
    result = parsing.parse_csv(csv_text.encode())
    got = [q.question for q in result.questions]
    assert got == ["Do you review user access periodically?", "Are audit logs retained?"], got
